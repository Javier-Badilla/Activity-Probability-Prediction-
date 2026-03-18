"""
=============================================================================
PLIP AFP Analysis Pipeline
=============================================================================
"""

import os
import re
import glob
import sys
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION — edit this section
# =============================================================================

# Folder containing one sub-folder per peptide, each with a PLIP report inside
PLIP_DIR = "./plip_out"

# Output folder and filename
OUTPUT_DIR  = "./plip_results"
OUTPUT_FILE = "PLIP_AFP_Summary.xlsx"

# Peptides to process (folder names inside PLIP_DIR)
# Remove or add peptides as needed
PEPTIDES = ["n_meme2", "n_meme3", "n_meme4", "n_meme4-2", "n_meme5"]

# Ice-Binding Surface residue map
# Keys = peptide names (must match folder names above)
# Values = set of (RESTYPE, RESNR) tuples
IBS_MAP = {
    "n_meme2":  {("LEU", 2),  ("ILE", 3),  ("ILE", 5), ("THR", 7),
                ("ALA", 8), ("LEU", 9), ("THR", 10)},
    "n_meme3":  {("ALA", 2),  ("THR", 4),  ("THR", 7),  ("GLY", 8),
                ("ILE", 9)},
    "n_meme4":  {("ILE", 1),  ("LEU", 4),  ("VAL", 5),  ("GLY", 6),
                ("VAL", 9)},
    "n_meme4-2":  {("ILE", 1),  ("ILE", 4),  ("VAL", 5),  ("GLY", 6),
                ("VAL", 9)},
    "n_meme5":  {("LEU", 2),  ("ALA", 3),  ("GLY", 5)}
}

# H-bond geometry thresholds for quality assessment
CLASH_DIST     = 2.0   # Å — H-A distance below this = likely rigid docking artifact
GOOD_ANGLE     = 130.0 # degrees — donor angle above this = good geometry
MARGINAL_DIST  = 3.80  # Å — D-A distance above this = marginal contact

# =============================================================================
# PARSING
# =============================================================================

def find_report(peptide_dir: str) -> str | None:
    """Find the PLIP report file inside a peptide's output folder."""
    for name in ("report_full.txt", "report.txt"):
        p = os.path.join(peptide_dir, name)
        if os.path.isfile(p):
            return p
    # Also search one level deeper (PLIP sometimes nests by PDB name)
    matches = glob.glob(os.path.join(peptide_dir, "**", "report*.txt"), recursive=True)
    if matches:
        return matches[0]
    return None


def parse_hbond_row(line: str) -> dict | None:
    """
    Parse a single data row from a PLIP H-bond table.
    Returns a dict or None if the line is not a valid data row.

    Expected format (pipe-delimited):
    | RESNR | RESTYPE | RESCHAIN | RESNR_LIG | RESTYPE_LIG | RESCHAIN_LIG |
    | SIDECHAIN | DIST_H-A | DIST_D-A | DON_ANGLE | PROTISDON | ... |
    """
    line = line.strip()
    if not line.startswith("|"):
        return None
    parts = [p.strip() for p in line.split("|")]
    parts = [p for p in parts if p]  # remove empty strings from split edges

    # Need at least: RESNR RESTYPE RESCHAIN RESNR_LIG RESTYPE_LIG RESCHAIN_LIG
    #                SIDECHAIN DIST_H-A DIST_D-A DON_ANGLE PROTISDON ...
    if len(parts) < 11:
        return None

    # Skip header rows
    if parts[0].strip() in ("RESNR", "======="):
        return None

    try:
        return {
            "RESNR":       int(parts[0]),
            "RESTYPE":     parts[1],
            "RESCHAIN":    parts[2],
            "RESNR_LIG":   int(parts[3]),
            "RESTYPE_LIG": parts[4],
            "SIDECHAIN":   parts[6].lower() == "true",
            "DIST_HA":     float(parts[7]),
            "DIST_DA":     float(parts[8]),
            "DON_ANGLE":   float(parts[9]),
            "PROTISDON":   parts[10].lower() == "true",
        }
    except (ValueError, IndexError):
        return None


def assess_contact(row: dict) -> str:
    """Assign a short quality label to a single H-bond."""
    dha   = row["DIST_HA"]
    dda   = row["DIST_DA"]
    angle = row["DON_ANGLE"]

    if dha < CLASH_DIST:
        return "Borderline (short)"
    if dda > MARGINAL_DIST:
        return "Marginal (long)"
    if angle >= GOOD_ANGLE:
        return "Good"
    if angle >= 110:
        return "Acceptable"
    return "Weak angle"


def parse_plip_report(report_path: str) -> list[dict]:
    """
    full PLIP report and return a list of H-bond dicts.
    Handles the per-ice-molecule block structure PLIP uses.
    Deduplicates rows that appear in overlapping blocks.
    """
    bonds = []
    seen  = set()

    with open(report_path, "r", encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()

    in_hbond_table = False
    for line in lines:
        stripped = line.strip()

        # Detect start of H-bond table
        if "**Hydrogen Bonds**" in stripped:
            in_hbond_table = True
            continue

        # Two blank lines or a new section header ends the table
        if in_hbond_table:
            if stripped == "" or stripped.startswith("**") or stripped.startswith("ICE:"):
                in_hbond_table = False
                continue

            row = parse_hbond_row(stripped)
            if row is None:
                continue

            # Deduplicate: same protein residue + same ligand molecule + same distances
            key = (row["RESNR"], row["RESTYPE"], row["RESNR_LIG"],
                   row["DIST_HA"], row["DIST_DA"])
            if key in seen:
                continue
            seen.add(key)
            bonds.append(row)

    return bonds


def classify_bonds(bonds: list[dict], ibs_set: set) -> list[dict]:
    """Add IBS classification and quality assessment to each bond."""
    result = []
    for b in bonds:
        is_ibs = (b["RESTYPE"], b["RESNR"]) in ibs_set
        b = dict(b)
        b["IS_IBS"]     = is_ibs
        b["IBS_LABEL"]  = "IBS" if is_ibs else "Non-IBS"
        b["ASSESSMENT"] = assess_contact(b)
        result.append(b)
    return result


def summarise(peptide: str, bonds: list[dict], ibs_set: set) -> dict:
    """summary statistics for one peptide."""
    ibs_bonds    = [b for b in bonds if b["IS_IBS"]]
    nonibs_bonds = [b for b in bonds if not b["IS_IBS"]]
    total        = len(bonds)
    ibs_count    = len(ibs_bonds)
    nonibs_count = len(nonibs_bonds)
    ratio        = ibs_count / total if total > 0 else 0.0

    # Active IBS residues (those that made at least one contact)
    active_ibs = sorted({(b["RESTYPE"], b["RESNR"]) for b in ibs_bonds},
                        key=lambda x: x[1])
    active_ibs_str = ", ".join(f"{r}{n}" for r, n in active_ibs) if active_ibs else "None"

    # Active non-IBS residues
    active_nonibs = sorted({(b["RESTYPE"], b["RESNR"]) for b in nonibs_bonds},
                           key=lambda x: x[1])
    active_nonibs_str = ", ".join(f"{r}{n}" for r, n in active_nonibs) if active_nonibs else "None"

    # Best single H-bond geometry (lowest dist_HA among non-clash bonds)
    valid = [b for b in bonds if b["DIST_HA"] >= CLASH_DIST]
    best_dha   = min((b["DIST_HA"]  for b in valid), default=None)
    best_angle = max((b["DON_ANGLE"] for b in valid), default=None)

    # Verdict
    if total == 0:
        verdict = "NO CONTACTS — check PLIP run"
    elif ratio >= 0.60:
        verdict = "PASS — IBS dominant"
    elif ratio >= 0.40:
        verdict = "BORDERLINE — advance"
    else:
        verdict = "FAIL — non-IBS dominant"

    return {
        "Peptide":              peptide,
        "Total H-bonds":        total,
        "IBS contacts":         ibs_count,
        "Non-IBS contacts":     nonibs_count,
        "IBS ratio":            round(ratio, 2),
        "Active IBS residues":  active_ibs_str,
        "Active non-IBS res.":  active_nonibs_str,
        "Best H-A dist (A)":    round(best_dha, 2)   if best_dha   is not None else "—",
        "Best angle (deg)":     round(best_angle, 1) if best_angle is not None else "—",
        "Verdict":              verdict,
    }

# =============================================================================
# EXCEL WRITING
# =============================================================================

C_DARK   = "1F3864"
C_MED    = "2E75B6"
C_GREEN  = "E2EFDA"
C_RED    = "FCE4D6"
C_YELLOW = "FFF2CC"
C_GREY   = "F2F2F2"
C_WHITE  = "FFFFFF"

VERDICT_COLORS = {
    "PASS":       ("E2EFDA", "375623"),
    "BORDERLINE": ("FFF2CC", "7F6000"),
    "FAIL":       ("FCE4D6", "C00000"),
    "NO CONTACTS":("F2F2F2", "595959"),
}


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def style(cell, text, bg=C_WHITE, bold=False, fg="222222", size=10,
          center=True, wrap=True):
    cell.value = text
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap)
    cell.border    = thin_border()


def hdr(cell, text, bg=C_DARK, size=9):
    style(cell, text, bg=bg, bold=True, fg="FFFFFF", size=size, center=True)


def apply_col_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb: Workbook, summaries: list[dict]):
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "AFP Peptide PLIP Analysis — Summary"
    t.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    sub.value = "Auto-generated by plip_afp_pipeline.py  |  IBS = Ice-Binding Surface residues"
    sub.font      = Font(name="Arial", italic=True, size=9, color="FFFFFF")
    sub.fill      = PatternFill("solid", start_color=C_MED)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 15

    ws.row_dimensions[3].height = 8  # spacer

    # Headers
    cols = list(summaries[0].keys())
    ws.row_dimensions[4].height = 32
    for c, col in enumerate(cols, 1):
        hdr(ws.cell(4, c), col.replace(" ", "\n"), bg=C_MED, size=9)

    # Data rows
    for i, row in enumerate(summaries):
        r = 5 + i
        ws.row_dimensions[r].height = 40
        bg = C_WHITE if i % 2 == 0 else C_GREY

        for c, col in enumerate(cols, 1):
            val = row[col]

            if col == "Verdict":
                # Color-code verdict cell
                vkey = next((k for k in VERDICT_COLORS if k in str(val)), "NO CONTACTS")
                vbg, vfg = VERDICT_COLORS[vkey]
                style(ws.cell(r, c), str(val), bg=vbg, bold=True, fg=vfg, size=10)

            elif col == "IBS contacts":
                style(ws.cell(r, c), val, bg=C_GREEN, bold=True, size=10)

            elif col == "Non-IBS contacts":
                style(ws.cell(r, c), val, bg=C_RED, size=10)

            elif col == "IBS ratio":
                cell = ws.cell(r, c)
                cell.value  = val
                cell.font   = Font(name="Arial", bold=True, size=10,
                                   color="375623" if val >= 0.5 else
                                         "7F6000" if val >= 0.35 else "C00000")
                cell.fill      = PatternFill("solid", start_color=bg)
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = thin_border()

            elif col == "Peptide":
                style(ws.cell(r, c), val, bg=bg, bold=True, size=11)

            else:
                style(ws.cell(r, c), val, bg=bg, center=True, size=10)

    apply_col_widths(ws, [9, 11, 11, 14, 9, 26, 26, 13, 13, 24])


def write_detail_sheet(wb: Workbook, all_bonds: dict[str, list[dict]]):
    ws = wb.create_sheet("PLIP Detail")

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "PLIP H-bond Detail — All Peptides"
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    detail_cols = ["Peptide", "Residue", "Res #", "IBS?",
                   "Dist H-A (A)", "Dist D-A (A)", "Angle (deg)",
                   "Sidechain?", "PROTISDON", "Assessment", "Notes"]
    ws.row_dimensions[2].height = 30
    for c, h in enumerate(detail_cols, 1):
        hdr(ws.cell(2, c), h, bg=C_MED, size=9)

    IBS_BG    = {"IBS": C_GREEN, "Non-IBS": C_RED}
    row_idx   = 3

    for peptide, bonds in all_bonds.items():
        if not bonds:
            # Write a placeholder row so the peptide appears
            ws.row_dimensions[row_idx].height = 16
            style(ws.cell(row_idx, 1), peptide,      bg=C_YELLOW, bold=True, size=9)
            style(ws.cell(row_idx, 2), "No contacts detected", bg=C_YELLOW,
                  center=False, size=9)
            for c in range(3, 12):
                style(ws.cell(row_idx, c), "", bg=C_YELLOW)
            row_idx += 1
            continue

        for b in bonds:
            ws.row_dimensions[row_idx].height = 16
            bg      = IBS_BG.get(b["IBS_LABEL"], C_WHITE)
            is_bold = b["IS_IBS"]

            # Notes column: flag clash-range distances
            notes = ""
            if b["DIST_HA"] < CLASH_DIST:
                notes = "Short — rigid docking artifact"
            elif b["DIST_DA"] > MARGINAL_DIST:
                notes = "Dist marginal"

            style(ws.cell(row_idx,  1), peptide,            bg=bg, bold=True,    size=9)
            style(ws.cell(row_idx,  2), b["RESTYPE"],        bg=bg, bold=is_bold, size=9)
            style(ws.cell(row_idx,  3), b["RESNR"],          bg=bg,               size=9)
            style(ws.cell(row_idx,  4), b["IBS_LABEL"],      bg=bg, bold=is_bold, size=9)
            style(ws.cell(row_idx,  5), round(b["DIST_HA"],  2), bg=bg, size=9)
            style(ws.cell(row_idx,  6), round(b["DIST_DA"],  2), bg=bg, size=9)
            style(ws.cell(row_idx,  7), round(b["DON_ANGLE"],1), bg=bg, size=9)
            style(ws.cell(row_idx,  8), "Yes" if b["SIDECHAIN"]  else "No", bg=bg, size=9)
            style(ws.cell(row_idx,  9), "Yes" if b["PROTISDON"]  else "No", bg=bg, size=9)
            style(ws.cell(row_idx, 10), b["ASSESSMENT"],      bg=bg, bold=is_bold, size=9)
            style(ws.cell(row_idx, 11), notes,                bg=bg, center=False, size=9)
            row_idx += 1

    apply_col_widths(ws, [9, 9, 7, 10, 12, 12, 12, 11, 11, 18, 28])


def write_ibs_map_sheet(wb: Workbook, ibs_map: dict, peptides: list[str]):
    ws = wb.create_sheet("IBS Map")

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "Ice-Binding Surface (IBS) Residue Map"
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for c, h in enumerate(["Peptide", "IBS Residues", "# IBS residues"], 1):
        hdr(ws.cell(2, c), h, bg=C_MED)
    ws.row_dimensions[2].height = 22

    for i, pep in enumerate(peptides):
        r = 3 + i
        ws.row_dimensions[r].height = 20
        ibs = sorted(ibs_map.get(pep, set()), key=lambda x: x[1])
        ibs_str = ", ".join(f"{res}{num}" for res, num in ibs) if ibs else "Not defined"
        bg = C_WHITE if i % 2 == 0 else C_GREY
        style(ws.cell(r, 1), pep,      bg=bg, bold=True, size=10)
        style(ws.cell(r, 2), ibs_str,  bg=C_GREEN if ibs else C_YELLOW,
              center=False, size=10)
        style(ws.cell(r, 3), len(ibs), bg=bg, size=10)

    apply_col_widths(ws, [10, 50, 16])


# =============================================================================
# MAIN
# =============================================================================

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_bonds:   dict[str, list[dict]] = {}
    all_summaries: list[dict]          = []
    errors: list[str]                  = []

    print("\n" + "="*60)
    print("  PLIP AFP Analysis Pipeline")
    print("="*60)

    for peptide in PEPTIDES:
        pep_dir = os.path.join(PLIP_DIR, peptide)

        if not os.path.isdir(pep_dir):
            msg = f"  [{peptide}] WARNING: folder not found — {pep_dir}"
            print(msg)
            errors.append(msg)
            all_bonds[peptide]     = []
            all_summaries.append(summarise(peptide, [], IBS_MAP.get(peptide, set())))
            continue

        report_path = find_report(pep_dir)
        if not report_path:
            msg = f"  [{peptide}] WARNING: no report.txt found in {pep_dir}"
            print(msg)
            errors.append(msg)
            all_bonds[peptide]     = []
            all_summaries.append(summarise(peptide, [], IBS_MAP.get(peptide, set())))
            continue

        print(f"  [{peptide}] Parsing: {report_path}")
        raw_bonds = parse_plip_report(report_path)

        ibs_set = IBS_MAP.get(peptide, set())
        if not ibs_set:
            print(f"  [{peptide}] WARNING: no IBS residues defined in IBS_MAP")

        bonds     = classify_bonds(raw_bonds, ibs_set)
        summary   = summarise(peptide, bonds, ibs_set)

        all_bonds[peptide] = bonds
        all_summaries.append(summary)

        print(f"           H-bonds: {summary['Total H-bonds']}  |  "
              f"IBS: {summary['IBS contacts']}  |  "
              f"Non-IBS: {summary['Non-IBS contacts']}  |  "
              f"Ratio: {summary['IBS ratio']:.2f}  |  "
              f"{summary['Verdict']}")

    # Write Excel
    out_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    wb = Workbook()
    write_summary_sheet(wb, all_summaries)
    write_detail_sheet(wb, all_bonds)
    write_ibs_map_sheet(wb, IBS_MAP, PEPTIDES)
    wb.save(out_path)

    print("\n" + "-"*60)
    if errors:
        print("  Completed with warnings:")
        for e in errors:
            print(f"    {e}")
    else:
        print("  All peptides processed successfully.")
    print(f"  Output: {out_path}")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
